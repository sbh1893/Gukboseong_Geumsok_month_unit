import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

# 페이지 기본 설정
st.set_page_config(page_title="월별 규격별 집계기", layout="wide")

st.title("📊 월별 & 규격별 수량/금액 집계")
st.markdown("""
파일을 업로드하면 **월별(Month)**로 묶고, 그 안에서 **규격**별로 합쳐서 
**총 수량**과 **총 합계금액**을 계산해 줍니다.
""")

# 1. 파일 업로드
uploaded_file = st.file_uploader("엑셀 또는 CSV 파일을 업로드하세요", type=['xlsx', 'csv'])

if uploaded_file is not None:
    st.info("파일을 분석하고 있습니다...")

    # 2. 데이터 읽기 함수 (캐싱 적용)
    @st.cache_data
    def load_data(file):
        file.seek(0)
        # 엑셀 시도
        try:
            return pd.read_excel(file, header=2), "Excel"
        except:
            pass
        
        # CSV 시도
        encodings = ['utf-8', 'cp949', 'euc-kr']
        for enc in encodings:
            try:
                file.seek(0)
                return pd.read_csv(file, header=2, encoding=enc), f"CSV({enc})"
            except:
                pass
        return None, "Fail"

    df, msg = load_data(uploaded_file)

    if df is not None:
        try:
            # ------------------------------------------------------
            # 3. 데이터 전처리
            # ------------------------------------------------------
            df.columns = df.columns.astype(str).str.strip()

            if '규 격' in df.columns:
                df = df[~df['규 격'].astype(str).str.contains('합계', na=False)]

            # 필수 컬럼 정의
            date_col = '납품일'
            spec_col = '규 격'
            qty_col = '수량'
            price_col = '합계금액'
            unit_col = '단위'

            if date_col in df.columns and spec_col in df.columns:
                # 데이터 채우기
                df[date_col] = df[date_col].ffill()
                df[spec_col] = df[spec_col].fillna("규격 미기재")
                
                # 날짜 변환 및 '월' 컬럼 생성
                df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                df['월'] = df[date_col].dt.strftime('%Y-%m')

                # 숫자 변환 (콤마 제거)
                for col in [qty_col, price_col]:
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.replace(',', '', regex=False)
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

                # ------------------------------------------------------
                # 4. 그룹화 (집계)
                # ------------------------------------------------------
                agg_dict = {}
                if qty_col in df.columns: agg_dict[qty_col] = 'sum'
                if price_col in df.columns: agg_dict[price_col] = 'sum'
                if unit_col in df.columns: agg_dict[unit_col] = 'first'

                # 월, 규격 기준으로 그룹화
                df_grouped = df.groupby(['월', spec_col]).agg(agg_dict)

                # 컬럼 순서 정리
                cols_order = []
                if unit_col in df_grouped.columns: cols_order.append(unit_col)
                if qty_col in df_grouped.columns: cols_order.append(qty_col)
                if price_col in df_grouped.columns: cols_order.append(price_col)
                
                df_final = df_grouped[cols_order]

                # 화면에 미리보기 출력
                st.success("집계 완료! 아래 결과가 엑셀로 저장됩니다.")
                st.dataframe(df_final)

                # ------------------------------------------------------
                # 5. 엑셀 생성 및 스타일링 (메모리 상에서 처리)
                # ------------------------------------------------------
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_final.to_excel(writer, sheet_name='Sheet1')

                # 스타일링 적용
                output.seek(0)
                wb = load_workbook(output)
                ws = wb.active

                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                center_align = Alignment(horizontal='center', vertical='center')
                right_align = Alignment(horizontal='right', vertical='center')

                # 숫자 컬럼 위치 찾기
                number_col_indices = []
                for cell in ws[1]:
                    if cell.value in [qty_col, price_col, '금액', '단 가']:
                        number_col_indices.append(cell.column)

                # 셀 스타일 적용
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                        
                        if cell.row == 1: # 헤더
                            cell.fill = header_fill
                            cell.font = Font(bold=True)
                            cell.alignment = center_align
                        else:
                            # 숫자 열은 우측 정렬 + 콤마, 나머지는 가운데
                            if cell.column in number_col_indices:
                                cell.alignment = right_align
                                cell.number_format = '#,##0'
                            else:
                                cell.alignment = center_align

                # 컬럼 너비 조정 (MergedCell 오류 방지 로직 적용)
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter # 첫 행은 항상 단일 셀이므로 안전
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_len:
                                max_len = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = (max_len + 2) * 1.2

                # 파일 저장을 위한 마무리
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                # 6. 다운로드 버튼
                st.download_button(
                    label="📥 엑셀 파일 다운로드",
                    data=output,
                    file_name="월별_규격별_집계표.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            else:
                st.error(f"필수 컬럼('{date_col}', '{spec_col}')을 찾을 수 없습니다.")
        except Exception as e:
            st.error(f"오류가 발생했습니다: {e}")
    else:
        st.error("파일을 읽을 수 없습니다. 형식을 확인해주세요.")